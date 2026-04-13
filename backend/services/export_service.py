"""Excel 数据导出服务"""
import io
from datetime import date, datetime
from typing import Optional, Dict, Any
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.work_order import WorkOrder
from models.milestone import Milestone
from models.progress import ProgressReport
from models.audit_log import AuditLog
from models.user import User

# ---------- 样式常量 ----------
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
CELL_FONT = Font(name="Microsoft YaHei", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")

STATUS_MAP = {
    "Backlog": "待办", "InProgress": "进行中", "Completed": "已完成",
    "OnHold": "暂停", "Cancelled": "已取消", "Pending": "待开始",
}

HEALTH_MAP = {
    "Green": "健康", "Yellow": "关注", "Red": "预警", "Gray": "未开始",
}


def _style_header(ws, row: int = 1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_cells(ws, start_row: int = 2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _fmt_date(d) -> str:
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d %H:%M")
    if isinstance(d, date):
        return d.strftime("%Y-%m-%d")
    return str(d)


def _status_label(s: str) -> str:
    return STATUS_MAP.get(s, s) if s else ""


def _health_label(h: str) -> str:
    return HEALTH_MAP.get(h, h) if h else ""


# ============ 导出函数 ============

def export_work_orders_to_excel(db: Session, filters: Dict[str, Any]) -> bytes:
    q = db.query(WorkOrder).order_by(WorkOrder.created_at.desc())
    if filters.get("status"):
        q = q.filter(WorkOrder.status == filters["status"])
    if filters.get("keyword"):
        kw = f"%{filters['keyword']}%"
        q = q.filter(
            (WorkOrder.wo_number.ilike(kw)) | (WorkOrder.project_name.ilike(kw))
        )
    if filters.get("is_delayed") is not None:
        q = q.filter(WorkOrder.is_delayed == filters["is_delayed"])

    items = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "工单列表"
    headers = ["工单号", "项目名称", "客户", "状态", "健康度", "总进度(%)", "计划交期", "创建时间"]
    ws.append(headers)
    _style_header(ws)

    for wo in items:
        ws.append([
            wo.wo_number, wo.project_name, wo.customer_name,
            _status_label(wo.status), _health_label(wo.health_status),
            wo.total_progress or 0,
            _fmt_date(wo.planned_delivery_date),
            _fmt_date(wo.created_at),
        ])

    _style_cells(ws)
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_work_order_detail_to_excel(db: Session, wo_id: int) -> bytes:
    wo = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not wo:
        raise ValueError(f"工单 {wo_id} 不存在")

    wb = Workbook()

    # Sheet 1: 基本信息
    ws1 = wb.active
    ws1.title = "基本信息"
    info = [
        ("工单号", wo.wo_number), ("项目名称", wo.project_name),
        ("客户", wo.customer_name), ("状态", _status_label(wo.status)),
        ("健康度", _health_label(wo.health_status)),
        ("总进度", f"{wo.total_progress or 0}%"),
        ("计划交期", _fmt_date(wo.planned_delivery_date)),
        ("创建时间", _fmt_date(wo.created_at)),
        ("更新时间", _fmt_date(wo.updated_at)),
    ]
    for row in info:
        ws1.append(row)
    _style_header(ws1)
    _auto_width(ws1)

    # Sheet 2: 里程碑
    ws2 = wb.create_sheet("里程碑进度")
    ms_headers = ["节点名称", "计划开始", "计划结束", "实际开始", "实际结束", "完成率(%)", "偏差天数", "状态"]
    ws2.append(ms_headers)
    _style_header(ws2)
    milestones = db.query(Milestone).filter(Milestone.wo_id == wo_id).order_by(Milestone.sort_order).all()
    for m in milestones:
        deviation = None
        if m.planned_end_date and m.actual_end_date:
            deviation = (m.actual_end_date - m.planned_end_date).days
        ws2.append([
            m.node_name, _fmt_date(m.planned_start_date), _fmt_date(m.planned_end_date),
            _fmt_date(m.actual_start_date), _fmt_date(m.actual_end_date),
            m.completion_rate or 0, deviation, _status_label(m.status),
        ])
    _style_cells(ws2)
    _auto_width(ws2)

    # Sheet 3: 进度汇报
    ws3 = wb.create_sheet("进度汇报")
    rp_headers = ["汇报日期", "汇报人", "班组", "完成率(%)", "备注"]
    ws3.append(rp_headers)
    _style_header(ws3)
    reports = db.query(ProgressReport).filter(ProgressReport.wo_id == wo_id).order_by(ProgressReport.report_date.desc()).all()
    for r in reports:
        ws3.append([
            _fmt_date(r.report_date), r.reported_by or "",
            r.team_name or "", r.completion_rate or 0, r.remark or "",
        ])
    _style_cells(ws3)
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_progress_to_excel(db: Session, filters: Dict[str, Any]) -> bytes:
    q = (
        db.query(WorkOrder, Milestone)
        .join(Milestone, WorkOrder.id == Milestone.wo_id)
        .order_by(WorkOrder.created_at.desc())
    )

    items = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "进度汇总"
    headers = ["工单号", "项目名称", "节点名称", "完成率(%)", "偏差天数", "状态"]
    ws.append(headers)
    _style_header(ws)

    for wo, m in items:
        deviation = None
        if m.planned_end_date and m.actual_end_date:
            deviation = (m.actual_end_date - m.planned_end_date).days
        ws.append([
            wo.wo_number, wo.project_name, m.node_name,
            m.completion_rate or 0, deviation, _status_label(m.status),
        ])

    _style_cells(ws)
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_audit_logs_to_excel(db: Session, filters: Dict[str, Any]) -> bytes:
    q = db.query(AuditLog).order_by(AuditLog.created_at.desc())
    if filters.get("user_id"):
        q = q.filter(AuditLog.user_id == filters["user_id"])
    if filters.get("action"):
        q = q.filter(AuditLog.action == filters["action"])
    if filters.get("resource_type"):
        q = q.filter(AuditLog.resource_type == filters["resource_type"])
    if filters.get("start_date"):
        q = q.filter(AuditLog.created_at >= filters["start_date"])
    if filters.get("end_date"):
        q = q.filter(AuditLog.created_at <= filters["end_date"])

    items = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "操作日志"
    headers = ["时间", "用户", "操作", "资源类型", "资源ID", "详情"]
    ws.append(headers)
    _style_header(ws)

    for log in items:
        ws.append([
            _fmt_date(log.created_at), log.username or "",
            log.action, log.resource_type, log.resource_id or "",
            log.detail or "",
        ])

    _style_cells(ws)
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_users_to_excel(db: Session) -> bytes:
    items = db.query(User).order_by(User.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "用户列表"
    headers = ["ID", "用户名", "姓名", "部门", "角色", "状态", "创建时间"]
    ws.append(headers)
    _style_header(ws)

    for u in items:
        ws.append([
            u.id, u.username, u.display_name or "",
            u.department or "", u.role or "",
            "在岗" if u.is_active else "离岗",
            _fmt_date(u.created_at),
        ])

    _style_cells(ws)
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
